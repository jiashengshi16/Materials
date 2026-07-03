import csv
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Hardcoded inputs / outputs. Put this script beside the two input files, or edit these paths.
ERRORS_CSV = Path("successful_run_errors.csv")
CANDIDATES_JSON = Path("material_similarity_candidates.json")
OUT_XLSX = Path("candidate_run_error_table.xlsx")
OUT_CSV = Path("candidate_run_error_table.csv")

# Set to True if you want a shorter run label in the spreadsheet/csv.
# False keeps the full run_id from successful_run_errors.csv.
SHORTEN_RUN_ID = False


def clean_num_wann(value):
    """Format num_wann consistently while preserving missing values."""
    if value is None or value == "":
        return ""
    try:
        return str(int(float(value)))
    except ValueError:
        return str(value)


def format_ratio(value):
    """Format ratios compactly but numerically enough for comparison."""
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):.6g}"
    except ValueError:
        return str(value)


def run_label(run_id):
    """Return either the full run_id or a shorter human-readable run label."""
    if not SHORTEN_RUN_ID:
        return run_id

    # Example run_id tail: ...__attempt_00089__num_wann_040__Ag2Hf2/Ag2Hf2__syitNxR
    attempt = re.search(r"attempt_\d+", run_id)
    tail = run_id.rsplit("/", 1)[-1]
    if attempt:
        return f"{attempt.group(0)} | {tail}"
    return tail


def read_inputs():
    with CANDIDATES_JSON.open("r", encoding="utf-8") as f:
        candidates_by_material = json.load(f)

    # material -> list of run records
    runs_by_material = defaultdict(list)
    # material -> all num_wann values seen for that material
    num_wann_by_material = defaultdict(set)

    with ERRORS_CSV.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        required = {"material", "run_id", "num_wann", "gemini_to_reference_ratio"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"{ERRORS_CSV} is missing required columns: {sorted(missing)}")

        for row in reader:
            material = row["material"]
            run_id = row["run_id"]
            num_wann = clean_num_wann(row.get("num_wann", ""))
            ratio = format_ratio(row.get("gemini_to_reference_ratio", ""))

            if num_wann:
                num_wann_by_material[material].add(num_wann)

            runs_by_material[material].append({
                "run_id": run_id,
                "run_label": run_label(run_id),
                "error_ratio": ratio,
                "num_wann": num_wann,
            })

    # Stable ordering: input order from CSV is usually meaningful, so preserve it.
    return candidates_by_material, runs_by_material, num_wann_by_material


def num_wann_text(material, num_wann_by_material):
    vals = sorted(num_wann_by_material.get(material, []), key=lambda x: (len(x), x))
    return ", ".join(vals)


def build_flat_rows(candidates_by_material, runs_by_material, num_wann_by_material):
    max_runs = 0
    flat_rows = []

    for material, candidates in candidates_by_material.items():
        material_num_wann = num_wann_text(material, num_wann_by_material)

        for candidate in candidates:
            candidate_runs = runs_by_material.get(candidate, [])
            max_runs = max(max_runs, len(candidate_runs))

            row = {
                "material": material,
                "material_num_wann": material_num_wann,
                "candidate_material": candidate,
                "candidate_num_wann": num_wann_text(candidate, num_wann_by_material),
                "runs": candidate_runs,
            }
            flat_rows.append(row)

    return flat_rows, max_runs


def write_csv(flat_rows, max_runs):
    headers = ["material", "material_num_wann", "candidate_material", "candidate_num_wann"]
    for i in range(1, max_runs + 1):
        headers += [f"run_{i}", f"error_ratio_{i}"]

    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()

        for row in flat_rows:
            out = {
                "material": row["material"],
                "material_num_wann": row["material_num_wann"],
                "candidate_material": row["candidate_material"],
                "candidate_num_wann": row["candidate_num_wann"],
            }
            for idx, run in enumerate(row["runs"], start=1):
                out[f"run_{idx}"] = run["run_label"]
                out[f"error_ratio_{idx}"] = run["error_ratio"]
            writer.writerow(out)


def write_xlsx(flat_rows, candidates_by_material, max_runs):
    wb = Workbook()
    ws = wb.active
    ws.title = "candidate_runs"

    headers = ["Material", "material_i num_wann", "Candidate material", "candidate num_wann"]
    for i in range(1, max_runs + 1):
        headers += [f"Run {i}", f"Error ratio {i}"]
    ws.append(headers)

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Write rows and remember group ranges for merged material cells.
    row_idx = 2
    group_ranges = []
    for material, candidates in candidates_by_material.items():
        start_row = row_idx
        group = [r for r in flat_rows if r["material"] == material]
        for row in group:
            values = [
                row["material"],
                row["material_num_wann"],
                row["candidate_material"],
                row["candidate_num_wann"],
            ]
            for run in row["runs"]:
                values += [run["run_label"], run["error_ratio"]]
            # Pad missing run columns.
            while len(values) < len(headers):
                values.append("")
            ws.append(values)
            row_idx += 1
        end_row = row_idx - 1
        if end_row >= start_row:
            group_ranges.append((start_row, end_row))

    # Merge material + material num_wann cells across each candidate block.
    for start_row, end_row in group_ranges:
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
        for col in (1, 2):
            ws.cell(start_row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(start_row, col).font = Font(bold=True)

    # Body styling.
    group_fill = PatternFill("solid", fgColor="EAF2F8")
    candidate_fill = PatternFill("solid", fgColor="F8FBFD")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[0].fill = group_fill
        row[1].fill = group_fill
        row[2].fill = candidate_fill
        row[3].fill = candidate_fill

    # Make ratio columns numeric where possible and right-aligned.
    for col in range(6, ws.max_column + 1, 2):
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, col)
            if cell.value not in (None, ""):
                try:
                    cell.value = float(cell.value)
                    cell.number_format = "0.0000"
                except (TypeError, ValueError):
                    pass
            cell.alignment = Alignment(horizontal="right", vertical="top")

    # Column widths: material/candidate readable; run IDs wrapped.
    widths = {
        1: 16,
        2: 16,
        3: 20,
        4: 18,
    }
    for col_idx in range(5, ws.max_column + 1):
        widths[col_idx] = 44 if (col_idx - 5) % 2 == 0 else 13
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions

    # Slightly taller rows so wrapped run IDs are visible.
    ws.row_dimensions[1].height = 28
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 42

    wb.save(OUT_XLSX)


def main():
    candidates_by_material, runs_by_material, num_wann_by_material = read_inputs()
    flat_rows, max_runs = build_flat_rows(candidates_by_material, runs_by_material, num_wann_by_material)

    write_csv(flat_rows, max_runs)
    write_xlsx(flat_rows, candidates_by_material, max_runs)

    print(f"Wrote {OUT_CSV} with {len(flat_rows)} candidate rows")
    print(f"Wrote {OUT_XLSX} with {len(flat_rows)} candidate rows")


if __name__ == "__main__":
    main()

