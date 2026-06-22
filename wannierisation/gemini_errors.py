#!/usr/bin/env python3

import json
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# =========================
# ONLY INPUT YOU EDIT
# =========================

INPUT_JSON = "/Users/jshi/Documents/GitHub/WannierisationBenchmarking/jobs/num_wann_ordered_diagnostics_summary.json"
OUTPUT_XLSX = "/Users/jshi/Documents/GitHub/WannierisationBenchmarking/jobs/gemini_vs_reference_errors.xlsx"

# Highlight rows where Gemini error is >= this many times the reference error.
RATIO_HIGHLIGHT_THRESHOLD = 10.0

# =========================


def is_number(x):
    return isinstance(x, (int, float)) and not isinstance(x, bool) and math.isfinite(x)


def get_first_number(d, keys):
    for key in keys:
        value = d.get(key)
        if is_number(value):
            return float(value)
    return None


def extract_rows(data):
    rows = []

    records = data.get("results", [])
    if not isinstance(records, list):
        records = []

    for d in records:
        if not isinstance(d, dict):
            continue

        # Only successful materials with nonzero reward.
        if d.get("successful") is not True:
            continue

        reward = d.get("reward")
        if not is_number(reward) or float(reward) == 0.0:
            continue

        material = d.get("material") or d.get("material_from_folder")

        # In your JSON, Gemini error is usually rmse_eV. Keep a fallback list
        # in case future summaries write it under a more explicit key.
        gemini_error = get_first_number(
            d,
            [
                "gemini_offmesh_rmse_eV",
                "rmse_eV",
                "offmesh_rmse_eV",
            ],
        )

        reference_error = get_first_number(
            d,
            [
                "reference_offmesh_rmse_eV",
                "reference_rmse_eV",
                "ref_offmesh_rmse_eV",
            ],
        )

        ratio = get_first_number(
            d,
            [
                "gemini_to_reference_rmse_ratio",
                "ratio",
            ],
        )

        if ratio is None and is_number(gemini_error) and is_number(reference_error) and reference_error > 0:
            ratio = gemini_error / reference_error

        rows.append(
            {
                "material": material,
                "num_wann": get_first_number(d, ["num_wann", "num_wann_from_folder", "num_target_bands"]),
                "reward": float(reward),
                "gemini_error_eV": gemini_error,
                "reference_error_eV": reference_error,
                "gemini_to_reference_ratio": ratio,
                "highlight_10x_or_more": bool(is_number(ratio) and ratio >= RATIO_HIGHLIGHT_THRESHOLD),
                "reference_error_source": d.get("reference_error_source"),
                "job_folder": d.get("job_folder"),
                "trial_folder": d.get("trial_folder"),
                "diagnostics_path": d.get("diagnostics_path"),
            }
        )

    return rows


def write_xlsx(rows, output_xlsx):
    wb = Workbook()
    ws = wb.active
    ws.title = "Gemini vs Reference"

    headers = [
        "material",
        "num_wann",
        "reward",
        "gemini_error_eV",
        "reference_error_eV",
        "gemini_to_reference_ratio",
        "highlight_10x_or_more",
        "reference_error_source",
        "job_folder",
        "trial_folder",
        "diagnostics_path",
    ]

    ws.append(headers)

    for row in rows:
        ws.append([row.get(h) for h in headers])

    # Styling
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(color="9C0006", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Number formats.
    number_formats = {
        "num_wann": "0",
        "reward": "0.000000",
        "gemini_error_eV": "0.000000E+00",
        "reference_error_eV": "0.000000E+00",
        "gemini_to_reference_ratio": "0.000",
    }

    header_to_col = {cell.value: cell.column for cell in ws[1]}
    ratio_col = header_to_col["gemini_to_reference_ratio"]

    for row_idx in range(2, ws.max_row + 1):
        ratio_value = ws.cell(row=row_idx, column=ratio_col).value
        highlight = is_number(ratio_value) and ratio_value >= RATIO_HIGHLIGHT_THRESHOLD

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical="top")

            header = ws.cell(row=1, column=col_idx).value
            if header in number_formats:
                cell.number_format = number_formats[header]

            if highlight:
                cell.fill = red_fill
                cell.font = red_font

    # Freeze header and add filters.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Column widths.
    width_by_header = {
        "material": 16,
        "num_wann": 10,
        "reward": 12,
        "gemini_error_eV": 18,
        "reference_error_eV": 20,
        "gemini_to_reference_ratio": 24,
        "highlight_10x_or_more": 22,
        "reference_error_source": 28,
        "job_folder": 58,
        "trial_folder": 24,
        "diagnostics_path": 80,
    }

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_by_header.get(header, 16)

    # Sort rows by ratio descending, if possible, before saving.
    # Since openpyxl does not physically sort, we sort before writing in main().

    output_xlsx = Path(output_xlsx).expanduser().resolve()
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    return output_xlsx


def main():
    input_json = Path(INPUT_JSON).expanduser().resolve()

    with open(input_json, "r") as f:
        data = json.load(f)

    rows = extract_rows(data)

    # Put worst Gemini/reference cases at the top.
    rows.sort(
        key=lambda r: r["gemini_to_reference_ratio"] if is_number(r.get("gemini_to_reference_ratio")) else -1.0,
        reverse=True,
    )

    if not rows:
        raise RuntimeError("No rows found with successful == true and nonzero reward.")

    output_xlsx = write_xlsx(rows, OUTPUT_XLSX)

    highlighted_count = sum(r["highlight_10x_or_more"] for r in rows)
    comparable_count = sum(is_number(r.get("gemini_to_reference_ratio")) for r in rows)

    print(f"Read: {input_json}")
    print(f"Wrote: {output_xlsx}")
    print(f"Rows written: {len(rows)}")
    print(f"Rows with comparable Gemini/reference ratio: {comparable_count}")
    print(f"Rows highlighted ratio >= {RATIO_HIGHLIGHT_THRESHOLD:g}: {highlighted_count}")


if __name__ == "__main__":
    main()
