#!/usr/bin/env python3
"""Add scalar Wannier-error diagnostics to the Gemini/reference workbook.

Only rows whose Gemini/reference RMSE ratio exceeds ``--ratio-threshold`` are
analysed. Array-valued diagnostics (RMSE by band, RMSE by k-point, full lists
of worst points, and .win parameter differences) intentionally are not put in
spreadsheet cells.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SUMMARY = ROOT / "jobs" / "num_wann_ordered_diagnostics_summary.json"
DEFAULT_INPUT_XLSX = ROOT / "jobs" / "gemini_vs_reference_errors.xlsx"
DEFAULT_OUTPUT_XLSX = ROOT / "jobs" / "gemini_vs_reference_errors_detailed.xlsx"
DEFAULT_DATASET = ROOT / "harbor_datasets" / "wannier_200"
SPREAD_RE = re.compile(
    r"Omega\s+(I|D|OD|Total)\s*=\s*([-+0-9.Ee]+)", re.IGNORECASE
)


SCALAR_COLUMNS = [
    "worst_abs_error_eV",
    "worst_kpoint_index_0based",
    "worst_dft_band_index_1based",
    "near_fermi_rmse_eV",
    "far_below_fermi_rmse_eV",
    "fraction_abs_error_gt_0.1eV",
    "fraction_abs_error_gt_0.25eV",
    "fraction_abs_error_gt_0.5eV",
    "fraction_abs_error_gt_1eV",
    "gemini_omega_I_A2",
    "gemini_omega_D_A2",
    "gemini_omega_OD_A2",
    "gemini_omega_total_A2",
    "reference_omega_I_A2",
    "reference_omega_D_A2",
    "reference_omega_OD_A2",
    "reference_omega_total_A2",
    "gemini_warning_count",
    "reference_warning_count",
    "gemini_disentanglement_converged",
    "reference_disentanglement_converged",
    "analysis_error",
]


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def read_hr(path: Path) -> tuple[np.ndarray, np.ndarray]:
    with path.open(encoding="utf-8") as handle:
        next(handle)
        num_wann = int(next(handle).strip())
        nrpts = int(next(handle).strip())
        degeneracies: list[int] = []
        while len(degeneracies) < nrpts:
            degeneracies.extend(int(value) for value in next(handle).split())
        r_vectors = np.zeros((nrpts, 3), dtype=int)
        hoppings = np.zeros((nrpts, num_wann, num_wann), dtype=np.complex128)
        for ir in range(nrpts):
            for _ in range(num_wann * num_wann):
                parts = next(handle).split()
                r_vectors[ir] = tuple(int(value) for value in parts[:3])
                m, n = int(parts[3]) - 1, int(parts[4]) - 1
                hoppings[ir, m, n] = float(parts[5]) + 1j * float(parts[6])
    weights = np.asarray(degeneracies[:nrpts], dtype=float)[:, None, None]
    return r_vectors, hoppings / weights


def interpolate_bands(
    kpoints: np.ndarray, r_vectors: np.ndarray, hoppings: np.ndarray
) -> np.ndarray:
    phases = np.exp(2j * np.pi * (kpoints @ r_vectors.T))
    h_of_k = np.einsum("kr,rmn->kmn", phases, hoppings, optimize=True)
    h_of_k = 0.5 * (h_of_k + np.conjugate(np.swapaxes(h_of_k, 1, 2)))
    return np.linalg.eigvalsh(h_of_k)


def first_file(paths: list[Path], description: str) -> Path:
    for path in paths:
        if path.is_file():
            return path
    raise FileNotFoundError(f"could not locate {description}")


def trial_dir(record: dict[str, Any], jobs_root: Path) -> Path:
    folder = jobs_root / str(record["job_folder"])
    named_trial = record.get("trial_folder")
    if isinstance(named_trial, str) and (folder / named_trial).is_dir():
        return folder / named_trial
    candidates = sorted(path for path in folder.iterdir() if path.is_dir())
    candidates = [path for path in candidates if (path / "verifier").is_dir()]
    if len(candidates) != 1:
        raise FileNotFoundError(f"could not identify one trial directory under {folder}")
    return candidates[0]


def gemini_artifact(trial: Path, suffix: str) -> Path:
    direct = sorted((trial / "artifacts").glob(f"attempt_*/*{suffix}"))
    nested = sorted((trial / "artifacts" / "artifacts").glob(f"attempt_*/*{suffix}"))
    return first_file([*direct, *nested], f"Gemini *{suffix} artifact")


def final_spreads(path: Path) -> dict[str, float | None]:
    values: dict[str, float] = {}
    for name, value in SPREAD_RE.findall(path.read_text(encoding="utf-8", errors="replace")):
        values[name.lower()] = float(value)
    return {
        "I": values.get("i"),
        "D": values.get("d"),
        "OD": values.get("od"),
        "total": values.get("total"),
    }


def wout_status(path: Path) -> tuple[int, bool | None]:
    text = path.read_text(encoding="utf-8", errors="replace")
    warning_count = sum("warning" in line.lower() for line in text.splitlines())
    lower = text.lower()
    if "disentanglement convergence criteria satisfied" in lower:
        converged: bool | None = True
    elif "maximum number of disentanglement iterations reached" in lower:
        converged = False
    else:
        converged = None
    return warning_count, converged


def finite_rmse(values: np.ndarray) -> float | None:
    return float(np.sqrt(np.mean(values**2))) if values.size else None


def analyse_record(
    record: dict[str, Any], jobs_root: Path, dataset: Path, near_window: float
) -> dict[str, Any]:
    material = str(record["material"])
    trial = trial_dir(record, jobs_root)
    hr_path = gemini_artifact(trial, "_hr.dat")
    gemini_wout = gemini_artifact(trial, ".wout")
    reference_root = dataset / material / "tests" / "reference"
    reference_wout = reference_root / "wannier" / "output" / "wannier90" / "aiida.wout"
    offmesh = reference_root / "dft" / "offmesh" / "reference"

    start = int(record["target_dft_band_start"])
    end = int(record["target_dft_band_end"])
    fermi = float(record["fermi_energy_eV"])
    dft = np.load(offmesh / "bands.npy")[:, start - 1 : end]
    kpoints = np.load(offmesh / "kpoints.npy")
    r_vectors, hoppings = read_hr(hr_path)
    wannier = interpolate_bands(kpoints, r_vectors, hoppings)[:, : dft.shape[1]]
    delta = wannier - dft
    below = dft <= fermi
    eligible = np.where(below, np.abs(delta), -np.inf)
    flat_index = int(np.argmax(eligible))
    worst_k, worst_relative_band = np.unravel_index(flat_index, eligible.shape)
    below_delta = delta[below]

    near_mask = below & (dft >= fermi - near_window)
    far_mask = dft < fermi - near_window
    gemini_spreads = final_spreads(gemini_wout)
    reference_spreads = final_spreads(reference_wout)
    gemini_warnings, gemini_converged = wout_status(gemini_wout)
    reference_warnings, reference_converged = wout_status(reference_wout)

    result: dict[str, Any] = {
        "worst_abs_error_eV": float(eligible[worst_k, worst_relative_band]),
        "worst_kpoint_index_0based": int(worst_k),
        "worst_dft_band_index_1based": int(start + worst_relative_band),
        "near_fermi_rmse_eV": finite_rmse(delta[near_mask]),
        "far_below_fermi_rmse_eV": finite_rmse(delta[far_mask]),
        "gemini_warning_count": gemini_warnings,
        "reference_warning_count": reference_warnings,
        "gemini_disentanglement_converged": gemini_converged,
        "reference_disentanglement_converged": reference_converged,
    }
    for threshold, label in [(0.1, "0.1"), (0.25, "0.25"), (0.5, "0.5"), (1.0, "1")]:
        result[f"fraction_abs_error_gt_{label}eV"] = float(
            np.mean(np.abs(below_delta) > threshold)
        )
    for prefix, spreads in [("gemini", gemini_spreads), ("reference", reference_spreads)]:
        for component in ("I", "D", "OD", "total"):
            result[f"{prefix}_omega_{component}_A2"] = spreads[component]
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--summary", type=Path, default=DEFAULT_SUMMARY)
    parser.add_argument("--input-xlsx", type=Path, default=DEFAULT_INPUT_XLSX)
    parser.add_argument("--output-xlsx", type=Path, default=DEFAULT_OUTPUT_XLSX)
    parser.add_argument("--dataset", type=Path, default=DEFAULT_DATASET)
    parser.add_argument("--jobs-root", type=Path, default=ROOT / "jobs")
    parser.add_argument("--ratio-threshold", type=float, default=2.0)
    parser.add_argument("--near-fermi-window-eV", type=float, default=1.0)
    args = parser.parse_args()
    if args.near_fermi_window_eV <= 0:
        raise SystemExit("--near-fermi-window-eV must be positive")

    summary = read_json(args.summary)
    records = {
        (item.get("material"), item.get("job_folder")): item
        for item in summary.get("results", [])
        if isinstance(item, dict)
    }
    workbook = load_workbook(args.input_xlsx)
    sheet = workbook["Gemini vs Reference"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    for required in ("material", "job_folder", "gemini_to_reference_ratio"):
        if required not in headers:
            raise SystemExit(f"input workbook is missing column: {required}")

    start_column = sheet.max_column + 1
    header_fill = PatternFill("solid", fgColor="548235")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for offset, name in enumerate(SCALAR_COLUMNS):
        cell = sheet.cell(1, start_column + offset, name)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
        sheet.column_dimensions[get_column_letter(cell.column)].width = max(16, min(32, len(name) + 2))

    analysed = 0
    failed = 0
    for row_number in range(2, sheet.max_row + 1):
        ratio = sheet.cell(row_number, headers["gemini_to_reference_ratio"]).value
        if not isinstance(ratio, (int, float)) or float(ratio) <= args.ratio_threshold:
            continue
        key = (
            sheet.cell(row_number, headers["material"]).value,
            sheet.cell(row_number, headers["job_folder"]).value,
        )
        record = records.get(key)
        try:
            if record is None:
                raise KeyError(f"summary record not found for {key}")
            values = analyse_record(
                record, args.jobs_root, args.dataset, args.near_fermi_window_eV
            )
            analysed += 1
        except Exception as exc:
            values = {"analysis_error": f"{type(exc).__name__}: {exc}"}
            failed += 1
        for offset, name in enumerate(SCALAR_COLUMNS):
            cell = sheet.cell(row_number, start_column + offset, values.get(name))
            cell.border = border
            cell.alignment = Alignment(vertical="top")
            if name.startswith("fraction_"):
                cell.number_format = "0.000%"
            elif name.endswith("_eV") or name.endswith("_A2"):
                cell.number_format = "0.000000E+00"

    args.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output_xlsx)
    print(f"Wrote: {args.output_xlsx}")
    print(f"Rows analysed (ratio > {args.ratio_threshold:g}): {analysed}")
    print(f"Rows with analysis errors: {failed}")
    print(f"Near-Fermi definition: {args.near_fermi_window_eV:g} eV below Fermi")


if __name__ == "__main__":
    main()
