#!/usr/bin/env python3
"""Run Gemini directly on prior-run Wannierisation evidence.

No argparse. Edit MATERIALS and GEMINI_BIN below, then run:

    scripts/run_gemini_self_debug_reviews.py
"""

from __future__ import annotations

import csv
import json
import shutil
import subprocess
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
HARBOR_DATASET_ROOT = Path(
    "/Users/jshi/Documents/GitHub/WannierisationBenchmarking/"
    "harbor_datasets/wannier_200"
)

# Hardcoded experiment controls.
MATERIALS = [
    "NNb",
    "S2Ta",
    'He',
    'FLi',
    'Br2V',
    "Ni4Zr4",
    'Si2Ta4',
    'Se4Tl4',
    'Na6O6',
    'C4O12Sr4',
    'Al4O8Zn2',
    'C2Cd2O6',
    'Al2Os',
    'Mo4S6'
]
MODEL = "gemini-3.1-pro-preview"
GEMINI_BIN = "gemini"
OUTPUT_ROOT = ROOT / "jobs" / "gemini_self_debug_reviews"
ORIGINAL_TASK_INSTRUCTIONS = ROOT / "instruction.md"


JOBS_ROOT = ROOT / "jobs"
DIAGNOSTICS_SUMMARY = JOBS_ROOT / "num_wann_ordered_diagnostics_summary.json"
ERROR_WORKBOOK = JOBS_ROOT / "gemini_vs_reference_errors.xlsx"
FAILURE_MODES_DIR = JOBS_ROOT / "gemini_failure_modes"

def find_original_task_instructions(material: str) -> Path:
    material_dir = HARBOR_DATASET_ROOT / material

    candidates = [
        material_dir / "instruction.md"
    ]

    for candidate in candidates:
        if candidate.is_file():
            return candidate

    matches = sorted(material_dir.glob("*instruction*.md"))
    if len(matches) == 1:
        return matches[0]

    matches = sorted(material_dir.glob("*instructions*.md"))
    if len(matches) == 1:
        return matches[0]

    raise SystemExit(
        f"Could not find original task instructions for {material} in {material_dir}. "
        "Expected one of: instruction.md, instructions.md, prompt.md, task.md, "
        "or a unique *instruction*.md file."
    )


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def clean_dir(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path)
    path.mkdir(parents=True)


def copy_file(src: Path, dst: Path) -> None:
    if not src.is_file():
        raise FileNotFoundError(src)
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def summary_by_material() -> dict[str, dict[str, Any]]:
    data = read_json(DIAGNOSTICS_SUMMARY)
    records: dict[str, dict[str, Any]] = {}
    for record in data.get("results", []):
        material = record.get("material") or record.get("material_from_folder")
        if isinstance(material, str):
            records[material] = record
    return records


def csv_by_material(path: Path) -> dict[str, dict[str, str]]:
    if not path.is_file():
        return {}
    with path.open(newline="", encoding="utf-8") as handle:
        return {row["material"]: row for row in csv.DictReader(handle) if row.get("material")}


def xlsx_by_material(path: Path) -> dict[str, dict[str, Any]]:
    if not path.is_file():
        return {}
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    records: dict[str, dict[str, Any]] = {}
    for row in rows:
        record = dict(zip(headers, row))
        material = record.get("material")
        if isinstance(material, str):
            records[material] = record
    return records


def resolve_trial(material: str, summary: dict[str, dict[str, Any]]) -> tuple[dict[str, Any], Path]:
    record = summary.get(material)
    if record is None:
        raise SystemExit(f"Material {material!r} is not present in {DIAGNOSTICS_SUMMARY}")
    job_folder = record.get("job_folder")
    trial_folder = record.get("trial_folder")
    if not isinstance(job_folder, str) or not isinstance(trial_folder, str):
        raise SystemExit(f"Material {material!r} is missing job_folder/trial_folder")
    trial = JOBS_ROOT / job_folder / trial_folder
    if not trial.is_dir():
        raise SystemExit(f"Trial directory does not exist for {material}: {trial}")
    return record, trial


def find_attempt_file(attempt: Path, material: str, suffix: str) -> Path:
    exact = attempt / f"{material}{suffix}"
    if exact.exists():
        return exact
    matches = sorted(attempt.glob(f"*{suffix}"))
    if len(matches) == 1:
        return matches[0]
    raise SystemExit(f"Could not resolve *{suffix} for {material} in {attempt}")


def prompt_text(material: str) -> str:
    return f"""# Gemini Self-Debug Review: {material}

You are reviewing a Wannierisation trajectory for `{material}`.
This is forensic analysis only. Do not rerun QE, do not produce a new
Wannierisation, and do not browse the internet. Use only files in this case
directory.

Your job is to reconstruct, as closely as the logs allow, the old decision chain 
using the original task instructions and old run logs. Treat 
`case_files/original_task_instructions.md` as the task prompt that was available 
to the old model during the original run.

Evaluate the trajectory fairly. If the old run made scientifically reasonable 
choices given the information available at the time, say so and do not force a 
critique. Only identify avoidable mistakes when the task materials, trajectory, 
logs, or final diagnostics provide evidence that a specific choice was poor, 
contradicted by later run output, or led to an avoidable failure or high RMSE. 
For any such issue, explain what a scientifically better second try would have 
changed, using only information that would have been available from the task 
materials and old run logs. 

Read these files:

- `case_files/diagnostics_record.json`
- `case_files/error_record.json`
- `case_files/artifacts/attempt_1/{material}.win`
- `case_files/artifacts/attempt_1/{material}.wout`
- `case_files/artifacts/attempt_1/run_manifest.json`
- `case_files/agent/trajectory.json`
- `case_files/agent/gemini-cli.trajectory.jsonl`
- `case_files/agent/gemini-cli.txt`
- `case_files/original_task_instructions.md`

Reference files and reference-derived fields are allowed only as outcome
comparators. Do not recommend "copy the reference",
SCDM, `use_ws_distance`, 'atom_proj = .true', or any other reference-only setting unless the original
task materials made that option available. If a reference setting is not
available under the original instructions, say so explicitly and give a
non-reference second-attempt change instead. Any proposed “better second try” must still obey 
case_files/original_task_instructions.md. Do not propose changes that would 
violate the original task prompt, including fixed num_bands, fixed target-band 
requirements, required artifact/status rules, forbidden external lookups, or
any other original instruction constraints.

Treat random projections as a bad/default-fallback choice unless the trajectory
gives unusually strong material-specific evidence. In general, random
projections should be discouraged. Do not say random projections were fine just
because Wannier90 completed. If random projections were used after the trajectory
had already derived a plausible physical projection, identify that abandoned
fork and explain why the fallback was avoidable.

Do not handwave from aggregate statistics. The core diagnosis must
come from this material's `.win`, `.wout`, run manifest, trajectory, and
diagnostics/error records.

Write exactly these two files:

- `self_debug_report.md`
- `self_debug_report.json`

The Markdown report must be step-by-step and specific. For each substantive
decision in the old trajectory, judge whether it was good, bad, mixed, or
uncertain, and explain why using concrete evidence from `.win`, `.wout`,
manifest notes, trajectory reasoning, and final error metrics. Cite the file
paths you used, and include line numbers when you have them from grep, rg, nl,
or similar inspection. Cover at least:

1. projection choice;
2. random projection use, if any;
3. `num_wann` / target-band handling;
4. `num_bands` / band-pool handling;
5. disentanglement outer and frozen windows;
6. response to Wannier90 warnings or iteration caps;
7. localization quality from WF spreads and spread components;
8. whether the old run accepted a result it should have rejected;

If the run shows evidence of avoidable issues, also cover:

9. the most likely specific failure chain;
10. what should be done differently next time.

For every decision review, answer all of these forensic questions:

- What exactly did the old run decide or claim?
- What evidence did the old run use at the time?
- What later evidence in `.wout`, diagnostics, errors, or trajectory contradicts
  or weakens that decision?
- Was the mistake avoidable without seeing the hidden reference recipe?
- What concrete second-attempt change should have been tried instead, staying
  WITHIN the original task instructions?
- What remains uncertain because the logs do not contain enough information?

Be especially suspicious of:

- accepting a result after checking only that `<seed>_hr.dat` exists;
- judging localization by average/total spread while ignoring max WF spread,
  spread outliers, final gradient, or iteration limits;
- declaring projections "excellent" or windows "robust" merely because they are
  chemically plausible;
- padding `num_wann` with duplicate same-site, same-angular-momentum
  projections without evidence that the channels are linearly independent;
- abandoning a physically motivated projection after a syntax, stale-file, or
  workflow error and falling back to `random`;
- recommendations that only say "increase iterations" when the logs show a
  projection, window, validation, or workflow-decision problem.

Be explicit about uncertainty. Do not claim causal proof when the files only
support diagnostic correlation. But make concrete judgments where evidence is
strong: random projections, unconverged disentanglement, huge WF spreads,
fragile windows, or mismatch between claimed rationale and observed output.
Use "plausible but unproven" rather than "good" when a choice is chemically
reasonable but the run output shows poor localization or band interpolation.
The goal is to find avoidable scientific decision errors, not to assign credit
for parameters that merely look conventional.

The JSON report must have this shape:

```json
{{
  "material": "{material}",
  "verdict": "good | mixed | bad | uncertain",
  "random_projections_used": true | false,
  "random_projection_verdict": "not_used | bad | uncertain",
  "decision_reviews": [
    {{
      "decision": "short name",
      "verdict": "good | mixed | bad | uncertain",
      "evidence": ["specific file-backed evidence"],
      "old_claim_or_decision": "what the old run said or did",
      "observed_failure_signal": "what later output showed",
      "why": "specific explanation",
      "better_choice": "concrete second-attempt change if the decision was an avoidable issue, otherwise null",
    }}
  ],
"failure_chain": ["ordered specific causes, or empty if no avoidable failure chain is supported"],
"recommended_next_run_changes": ["specific changes, or empty if the trajectory was reasonable and no supported changes are warranted"],
}}
```

Your final response should be a concise JSON object pointing to
`self_debug_report.md` and `self_debug_report.json`.
"""

def build_case(
    material: str,
    summary: dict[str, dict[str, Any]],
    errors: dict[str, dict[str, Any]],
    modes: dict[str, dict[str, str]],
) -> Path:
    record, trial = resolve_trial(material, summary)
    attempt = trial / "artifacts" / "attempt_1"
    if not attempt.is_dir():
        raise SystemExit(f"Missing attempt_1 for {material}: {attempt}")

    case_dir = OUTPUT_ROOT / material
    clean_dir(case_dir)
    case_files = case_dir / "case_files"

    copy_file(find_attempt_file(attempt, material, ".win"), case_files / "artifacts" / "attempt_1" / f"{material}.win")
    copy_file(find_attempt_file(attempt, material, ".wout"), case_files / "artifacts" / "attempt_1" / f"{material}.wout")
    copy_file(attempt / "run_manifest.json", case_files / "artifacts" / "attempt_1" / "run_manifest.json")
    copy_file(trial / "agent" / "trajectory.json", case_files / "agent" / "trajectory.json")

    for optional in ("gemini-cli.trajectory.jsonl", "gemini-cli.txt"):
        src = trial / "agent" / optional
        if src.exists():
            copy_file(src, case_files / "agent" / optional)

    write_text(case_files / "diagnostics_record.json", json.dumps(record, indent=2) + "\n")
    write_text(case_files / "error_record.json", json.dumps(errors.get(material, {}), indent=2) + "\n")
    write_text(case_files / "failure_modes_record.json", json.dumps(modes.get(material, {}), indent=2) + "\n")

    original_task_instructions = find_original_task_instructions(material)
    copy_file(
        original_task_instructions,
        case_files / "original_task_instructions.md",
    )

    write_text(case_dir / "prompt.md", prompt_text(material))
    return case_dir

def report_is_nonempty(case_dir: Path) -> bool:
    md_path = case_dir / "self_debug_report.md"
    json_path = case_dir / "self_debug_report.json"

    if not md_path.is_file() or not json_path.is_file():
        return False

    if not md_path.read_text(encoding="utf-8").strip():
        return False

    raw_json = json_path.read_text(encoding="utf-8").strip()
    if not raw_json:
        return False

    try:
        data = json.loads(raw_json)
    except json.JSONDecodeError:
        return False

    # Minimal sanity checks that this is actually the requested diagnosis.
    if not isinstance(data, dict):
        return False
    if not data.get("verdict"):
        return False
    if not data.get("decision_reviews"):
        return False
    if not data.get("failure_chain"):
        return False
    if not data.get("recommended_next_run_changes"):
        return False

    return True


def run_gemini(case_dir: Path) -> None:
    prompt = (case_dir / "prompt.md").read_text(encoding="utf-8")
    combined_log_path = case_dir / "gemini_stdout_stderr.txt"
    status_path = case_dir / "run_status.json"

    command = [
        GEMINI_BIN,
        "--yolo",
        f"--model={MODEL}",
        f"--prompt={prompt}",
    ]

    attempts: list[dict[str, Any]] = []
    max_attempts = 10

    for attempt_index in range(1, max_attempts + 1):
        # Remove stale outputs so success must come from this attempt.
        for output_name in ("self_debug_report.md", "self_debug_report.json"):
            output_path = case_dir / output_name
            if output_path.exists():
                output_path.unlink()

        attempt_log_path = case_dir / f"gemini_attempt_{attempt_index:02d}_stdout_stderr.txt"

        completed = subprocess.run(
            command,
            cwd=case_dir,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
        )

        stdout_stderr = completed.stdout or ""
        attempt_log_path.write_text(stdout_stderr, encoding="utf-8")

        produced_nonempty_diagnosis = report_is_nonempty(case_dir)

        attempt_record = {
            "attempt": attempt_index,
            "returncode": completed.returncode,
            "attempt_log_path": attempt_log_path.name,
            "self_debug_report_md_exists": (case_dir / "self_debug_report.md").is_file(),
            "self_debug_report_json_exists": (case_dir / "self_debug_report.json").is_file(),
            "produced_nonempty_diagnosis": produced_nonempty_diagnosis,
        }
        attempts.append(attempt_record)

        status = {
            "command": command,
            "max_attempts": max_attempts,
            "success": produced_nonempty_diagnosis,
            "attempts": attempts,
        }
        status_path.write_text(json.dumps(status, indent=2) + "\n", encoding="utf-8")

        if produced_nonempty_diagnosis:
            combined_log_path.write_text(stdout_stderr, encoding="utf-8")
            return

        print(
            f"Gemini attempt {attempt_index}/{max_attempts} did not produce "
            f"a non-empty diagnosis for {case_dir.name}; retrying..."
        )

    combined_log_path.write_text(
        "\n\n".join(
            [
                f"===== ATTEMPT {record['attempt']} "
                f"returncode={record['returncode']} =====\n"
                f"{(case_dir / record['attempt_log_path']).read_text(encoding='utf-8')}"
                for record in attempts
            ]
        ),
        encoding="utf-8",
    )

    raise SystemExit(
        f"Gemini review failed after {max_attempts} attempts or did not write "
        f"a non-empty diagnosis: {case_dir}"
    )


def main() -> None:
    if shutil.which(GEMINI_BIN) is None:
        raise SystemExit(
            f"Could not find {GEMINI_BIN!r} on PATH. Edit GEMINI_BIN at the top of this script "
            "or run it in the same environment where Gemini CLI is installed."
        )

    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    summary = summary_by_material()
    errors = xlsx_by_material(ERROR_WORKBOOK)
    modes = csv_by_material(FAILURE_MODES_DIR / "failure_modes.csv")

    for material in MATERIALS:
        case_dir = build_case(material, summary, errors, modes)
        print(f"Running Gemini review for {material} in {case_dir}")
        run_gemini(case_dir)
        print(f"Wrote expected outputs in {case_dir}")


if __name__ == "__main__":
    main()
